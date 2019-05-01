#######################################################################################################################
'''
Examples of Python Code Samples; examples.py
Contents:
  Lesson 01 - Print and Input commands
  Lesson 02 - Strings, variables
  Lesson 03 - Numeric data types, Operators, Type conversion, String methods
  Lesson 04 - Importing Modules, Conditions, Conditional branching, code blocks
  Lesson 05 - while loop, option menu, indexes
  Lesson 06 - for loops, range(), sequences, slices
  Lesson 07 - tuples, lists, nested sequences
  Lesson 08 - list sorting, itemgetter(), dictionaries
  Lesson 09 - Function basics, return values
  Lesson 10 - Function parameters, Formatting output, Looping inside an outer loop
  Lesson 11 - function scoping\global variables
  Lesson 12 - Exceptions, Python binary files, .csv files, Excel files
  '''


#######################################################################################################################
## Lesson 12
# =============================================================
# Exceptions
# =============================================================
vf_num = 0.0;
try:
    vf_num = float(input("\nEnter a number: ")); # Verify a numeric value
    print("Number you entered: ", vf_num);
except Exception as e:
    print("Data entry error - " + str(e));      # Print error message
    


# =============================================================
# Python Binary File Input\Output
# =============================================================
import pickle;
l_list = [ [ 'a', 2, 'apple' ] 
          ,[ 'b', 4, 'bravo' ] 
          ,[ 'c', 5, 'crate' ] 
         ];
d_dict = {};            # Initialize an empty dictionary
vs_file = 'data.dat';   # Filename
for i in l_list: d_dict[i[0]] = [float(i[1]), i[2]]; # Create the key i[0] and values list
print (l_list, d_dict);


# Write the data structures
try:
    f = open(vs_file, "wb");
    pickle.dump(l_list, f);
    pickle.dump(d_dict, f);
    f.close();
    print("Pickle File write successful");
except Exception as e:
    print("Pickle File Write Error: " + str(e));    # Print error message


# Read the data structures by un-pickling IN THE SAME ORDER AS WRITING
try:
    f = open(vs_file, "rb");
    l2_list = pickle.load(f);
    d2_dict = pickle.load(f);
    f.close();
    print("Pickle File read successful");
except Exception as e:
    print("Pickle File Read Error: " + str(e));    # Print error message


# Compare the data structures
if l2_list == l_list: print("Lists are equivalent");
if d2_dict == d_dict: print("Dictionaries are equivalent");    




# =============================================================
# .csv Files
# =============================================================
# *******************
# Reading from products.csv using csv.DictReader
# *******************
import csv;        # Load the library module
lst   = [];        # Initialize a list for contents of file


try:
    file   = open('products.csv', newline='');   # Open the file – default for reading
    reader = csv.DictReader(file);      # Return a dictionary reader iterator for the file
    for d_row in reader:                # Loop on each dictionary row in the file into a list
      # File title line:  prod_nbr, price, size, color
        l_items = [d_row['prod_nbr'],float(d_row['price']),d_row['size'],d_row['color']]; 
        lst.append(l_items);        # Append the next list row to a list of lists
    file.close();                  # close the file
    if not lst: print("Error - File is empty"); # Check if file\list is empty
    else:       print("File Read access successful");
except Exception as e:
    print("File processing error - " + str(e)); # Print error message


for i in lst: print(i); # Display file contents
input("DictReader method - Press Enter key");


# *******************
# Reading from orders.csv using csv.reader() method
# *******************
import csv;        # Load the library module
lst   = [];        # Initialize a list for contents of file
r_cnt = 0;        # File row counter


try:
    file = open('orders.csv', newline=''); # Open the file – default for reading
    reader = csv.reader(file);                # Return a reader iterator for the file
    for l_row in reader:                # Loop on each row in the file into a list
        if r_cnt == 0: r_cnt +=1; continue; # Skip 1st row containing column titles
        lst.append(l_row);        # Append the next list row to the list
    file.close();                  # close the file
    if not lst: print("Error - File is empty"); # Check if file\list is empty
    else:       print("File Read access successful");
except Exception as e:
    print("File processing error - " + str(e));    # Print error message


for i in lst: print(i); # Display file contents
input("Reader method - Press Enter key");




# *******************
# Write the file using the csv.writerow() method
# *******************
import csv;        # Load the library module
lst = [['child-lw',30,'2016-07-27'], ['adult-mg',12,'2016-10-03'], ['adult-mw',12,'2016-08-30']]; # Initialize a list for contents of file
try:
    file = open('ord_out.csv', 'w', newline=''); # Open the file for writing w\o '\n'
    writer = csv.writer(file);                   # Return a writer object for the file
    writer.writerow(('prod_nbr','qty','date'));  # Write the column titles in the first row
    for l_items in lst:                    # Loop on each list row in the list of lists
        writer.writerow(l_items);   # Write each list row
    file.close();                      # close the file
    print("File Write access successful");
except Exception as e:
    print("File processing error - " + str(e));    # Print error message




# =============================================================
# Excel Files 
# =============================================================
# *******************
# Get pip and install openpyxl module
# *******************
'''
# Load Pip to Load openpyxl
from os import system;  # System shell commands
system("python -m pip install --upgrade pip");
system("pip install openpyxl");
'''


# *******************
# Read from the Excel file 
# *******************
import openpyxl;            # for Excel workbook modules
wb_file = 'Excel_File.xlsx';    # Excel filename
ws_names = [];                  # List of worksheet names


# Open the workbook file
try:
    wb = openpyxl.load_workbook(filename = wb_file);  # Get the workbook object
except Exception as e:
    print("Excel file - ", wb_file, " Read Error: " + str(e));  # Print error message
    input("Press Enter to continue");


# Loop through the worksheet names in the workbook
for ws_name in wb.sheetnames:   # For each work sheet name in the workbook
    input("\nProcessing worksheet: " +ws_name);
    ws = wb[ws_name];           # Get the the worksheet object


  # Get an interator for all rows and columns in the worksheet object
    rows = ws.iter_rows(min_col = 1, min_row = 1, max_col = ws.max_column, max_row = ws.max_row)


  # Create List of Lists of rows and columns in the worksheet
    l_ws = [[cell.value for cell in list(row)] for row in rows];    
 
  # Remove title row if desired
    l_ws.pop(0);


  # Copy the worksheet list
    if   ws_name == 'Products':
        l_prod = l_ws;  # Create l_prod list
        print("\nWorksheet: ", ws_name);
        for i in l_prod: print(i);
        
    elif ws_name == 'Orders':
        l_ord  = l_ws;  # Create l_ord list
        print("\nWorksheet: ", ws_name);
        for i in l_ord: print(i);


    else: pass;     # Ignore any other worksheet


# Close excel file
try:
    wb.close(); # Close the workbook
except Exception as e:
    print("Excel file - ", wb_file, " Error on closing: " + str(e));  # Print error message
    input("Press Enter to continue");




# *******************
# Write to the Excel file 
# *******************
import openpyxl;             # Excel workbook modules
wb_file = 'Excel_File.xlsx'; # Excel filename
ws_sheet  = "Data"; # New worksheet to create


## -------------------------
## Open excel file
try:
    wb = openpyxl.load_workbook(wb_file);  # Open the workbook
except Exception as e:
    print("Excel file - ", wb_file, " Read Error: " + str(e));  # Print error message
    input("Press Enter to continue");


## -------------------------
## Process
if ws_sheet in wb.sheetnames:
    ws = wb[ws_sheet];  # Get worksheet object for worksheet name
    wb.remove(ws);      # Remove worksheet from file
wb.create_sheet(ws_sheet);
ws = wb[ws_sheet];      # Get worksheet object for worksheet name


## -------------------------
## Write data
ws.append(["One", "Two", "Three"]); # Title row
ws.append([1, 2, 3]);   # Rows can also be appended


## -------------------------
## Close excel file
try:
    wb.save(wb_file);
    wb.close(); # Close the workbook
    print("Worksheet:", ws_sheet, "added to file:", wb_file);
except Exception as e:
    print("Excel file - ", wb_file, " Error on closing: " + str(e));  # Print error message
    input("Press Enter to continue");






#######################################################################################################################
## Lesson 11
# =============================================================
# Function scope
# =============================================================
# ----------------------------
# Global Scope - reading a variable
# ----------------------------
def f1():
    x = 11;                            # variable named x is defined in Local scope
    print("Inside f1: x = ", x);
    print("Inside f1: y = ", y);    # Visible in f1 by declaration in global scope
    return; 


x = 2;                # variable named x is in Global scope        
y = 22;         # variable named y is in Global scope        
f1();        
print("Outside f1: x = ", x);
print("Outside f1: y = ", y);


# ----------------------------
# Function to Function scope - reading a Global variable
# ----------------------------
def f1():
    print("Inside f1: x = ", x);    # variable named x is in Global scope 
    #print("Inside f1: y = ", y);   # Error - variable y is NOT in Global scope 
    return;


def main():
    global x;        # x is in Global scope
    x = 2; 
    y = 22;     # variable named y is in local scope of main()
    f1();        # call f1() from inside main()
    return;


main();        # call main()




# =============================================================
# Passing Parameters into a function for modification
# =============================================================
# ----------------------------------
# Pass by value – Immutable data type
# ----------------------------------
def f1(p):
        p += 1;                                # Modifying p in the function
        print("Inside f1: p = ", p);    # Inside f1:  p =  2
        return;
x = 1;
f1(x);                        
print("Outside f1: x = ", x);          # Outside f1: x =  1


# ----------------------------------
# Pass by Reference – Mutable data type modified in the function
# ----------------------------------
def f1(p_list):        
        print("Inside f1 before modification = ", p_list);
        p_list += [4, 5, 6];
        print("Inside f1 after  modification = ", p_list);
        return;
list1 = [1, 2, 3];         
f1(list1);
print("Outside f1 = ", list1); # => Outside f1 =  [1, 2, 3, 4, 5, 6]


# ----------------------------------
# Pass by Reference – Mutable data type replaced in the function
# ----------------------------------
def f1(p_list):        
        print("Inside f1 before replace = ", p_list);
        p_list = [4, 5, 6];
        print("Inside f1 after replace = ", p_list);
        return;
list1 = [1, 2, 3];
f1(list1);
print("Outside f1 = ", list1);         # => Outside f1 =  [1, 2, 3]






#######################################################################################################################
## Lesson 10
# =============================================================
# Function input\output
# =============================================================
# --------------------------------------
# Function definitions


# Function with no Return values
def fn_0(): 
        print("In function fn_0"); 
        return;


# Function with Return values
def fn_1(): 
        vs_return = 'fn_1';
        print ("In function fn_1, returning: ", vs_return);
        return vs_return; 


## Function with Parameter
def fn_2(msg): 
        print ("In function fn_2, input parameter: ", msg); 
        return;


## Function with Parameters and Return values
def fn_3(n,msg): 
        vs_return = 'complete';
        print ("In function fn_3, input parameters: ", n, msg); 
        print ("In function fn_3, returning: ", vs_return);
        return (vs_return, n+1); 


# --------------------------------------
# Function calls
fn_0();              # Call function 0


vs_msg1 = fn_1();    # WARNING: Ensure the return variable is the same data type as the return value 
print(vs_msg1); 


fn_2('Parameter_21'); # Call function 1


vs_msg3, vi_x = fn_3(3, 'parameter_3');    # WARNING: Ensure the return variables is the same data type as the return value AND in the same order.
print(vs_msg3, vi_x);




# ================================
# Formatting Output
# ================================
# Get products list
from Project_Data import l_prod;


# Set the report header and data format
#           Product,  Price, Size,   Color
v_hdr_fmt = "|{0:^10}|{1:^7}|{2:^8}|{3:^8}|";  # Header format
v_dat_fmt = "|{0:<10}|{1:>7}|{2:<8}|{3:<8}|";  # Data   format


# Display the report header
print (v_hdr_fmt.format('----------', '------', '--------', '--------'));   # Header data
print (v_hdr_fmt.format('Product'   , 'Price' , 'Size'    , 'Color'   ));   # Header data
print (v_hdr_fmt.format('----------', '------', '--------', '--------'));   # Header data


# Display the data
import operator;    ## Module for itemgetter
for i in sorted(l_prod, key=operator.itemgetter(1), reverse=True):
    print (v_dat_fmt.format(i[0], '${:,.2f}'.format(i[4]), i[2], i[3]));




# =============================================================
# Looping inside an outer loop
# ======================================
# Initialize variables
import copy;
d = { 'a':['aa', '11'],  'b':['bb', '22'],  'c':['cc', '33'] }; # Dictionary of lists
lst = [ ['x', 1, 11.0],  ['y', 2, 22.0],  ['z', 3, 33.0] ];     # List of lists
out_lst = [];  # Output list to display


for k, v in d.items():  # Outer loop
   count = 0;           # Initialie outer loop counter


   for i in lst:        # Inner loop
      count += 1;       # Increment counter


   out_lst.append([k,v[0],v[1], count]);  # Append to output list key, values, couter


print(out_lst);   # Display output list


############################################################################################################
## Lesson 09
# =============================================================
# Basic Function syntax
# =============================================================
# **************************************
# No Return value
# **************************************
def f_err_msg():
    print(" Application Error !!");
    return;        # End of function f_err_msg


# Main program
vb_err = True;  # Set error status
if vb_err: f_err_msg();   # Call the function


# **************************************
# One Return value
# **************************************
def f_menu():
    choice = '';    # Initialize variable
    while choice not in ('1', '2', 'x', 'X'):
        print('''
        Menu
        1. Option 1
        2. Option 2
        x  Exit
              ''');
        choice = input("Enter Selection ");
    return choice.upper(); # Return selection option


# Main program
vs_opt = f_menu();  # Get menu selection 
print("Menu Selection: ", vs_opt);




# **************************************
# Multiple Return values
# **************************************
def f_login():
    f_user = input("Enter Username ");
    f_pass = input("Enter Password ");
    f_acct = input("Enter Account Number ");
    return f_user, f_pass, f_acct;      # Return login values


# Main program
vs_user, vs_pass, vs_acct = f_login();  # Get login values


print("Login User, Pass, Account: ", vs_user, vs_pass, vs_acct);




# =============================================================
# Using main() and doc_strings
# =============================================================
##############################################################
def f_name():
    '''  
    Function: f_name()
    This function is an example of basic syntax
    Inputs: None
    Outputs: None
    ''' 
 ## Local Variables
    vs_var = "f_name";        # Initialize local variable


    print("Inside f_name - vs_var: ", vs_var)    
    return;        # End of function f_name


##############################################################
def main(): 
    '''  
    Function: main()
    Main is used to control the overall processing
    Inputs: None
    Outputs: None
    ''' 


 ## Local Variables
    vs_var = "main";    # Initialize local variable
    print("Inside main - vs_var: ", vs_var)    


    print("Calling f_name:");
    f_name();
    print("\nBack inside the main function.");


    return; 


main(); # calling main()
print ("Documentation String for main(): ", main.__doc__);
print ("Documentation String for f_name(): ",f_name.__doc__);




############################################################################################################
## Lesson 08
# =============================================================
# List sorting and List of Lists
# =============================================================
# ------------------
# Sort METHOD -> the elements of a list sorted IN PLACE
l_list = ['apple', 'dog', 'echo', 'bravo', 'cat'];
print("List before sorting: ", l_list); 


l_list.sort();        # sort default – ascending order
print("List after ascending sort: ", l_list); 


l_list.sort(reverse=True);  # sort default – descending order
print("List after descending sort: ", l_list); 


# ------------------
# Sort FUNCTION -> the elements of a list sorted and returned as a new list
l_list = ['apple', 'dog', 'echo', 'bravo', 'cat'];
print("List before sorting: ", l_list); 


l_list2 = sorted(l_list);        # sort default – ascending order
print("NEW list after ascending sort: ", l_list2); 
print();




# ------------------
# Using the key function and just printing the returned list
l_list = ['ccc', 'aaaa', 'd', 'bb'];  # list
print("List before sort: ", l_list);
print("After sort on element length: ", sorted(l_list, key = len, reverse = False)); # sort on string len


# ------------------
# Sorting list of lists using itemgetter()
import operator;              # import the operator module
l_list = [ [ 1, 3, 3 ] 
          ,[ 2, 4, 6 ] 
          ,[ 3, 5, 7 ] 
         ];


# Sort method - sort the list in place in a specific order - Descending
l_list.sort(key=operator.itemgetter(2), reverse=True);
print(l_list);


# Sorted function - Sorting list of list on a specific column - Ascending
ls = sorted(l_list, key=operator.itemgetter(1), reverse=False);
print ("New list sorted ascending on element 1: ", ls);
print ("Original list: ", l_list);


# Looping through a sorted list of lists
for i in sorted(l_list, key=operator.itemgetter(1), reverse=False):
    print(i, i[0], i[1], i[2]);




# =============================================================
# Basic Dictionary operations
# =============================================================
d  = {'a':1, 'b':2, 'c':3};     # Initialize a dictionary
print (d['c'], type(d['a']));   # Access the value using the key
print (d['x']);  # ERROR if key does not exist


# Methods 
i = d.get('a');         # Get a value that exists
j = d.get('x');         # Get a value that does not exist
print(i, type(i), j, type(j));  # NO ERROR message if key ('x') does not exist


print(('a' in d), ('x' in d));  # Membership - Verify the key exists
print(len(d));                  # Count of key\value pairs


for k in d.keys():     print(k);   # keys()   returns the dictionary keys 
for v in d.values():   print(v);   # values() returns the dictionary values
for k, v in d.items(): print(k,v); # items()  returns the dictionary key and value


# Dictionary modification
d['x'] = 6;        # Add key:value pair = 'x':6 
print(d);


d['x'] = 999;   # Update value for key = 'x'
print(d);       # Warning: The syntax for add\update is identical 


del(d['x'])     # Delete the key x together with his value.
print(d);


# =============================================================
# Dictionary nested sequences
# =============================================================
# ------------------------------------------------
d = { 'a':[ 1, 2, 3 ] 
     ,'b':[ 2, 4, 6 ] 
     ,'c':[ 3, 5, 7 ] 
    };      # Dictionary of lists
print(d);   


print(d['a'], type(d['a']));    # To access a specific list, use the key 
print(d['a'][2]);               # To access a specific list element


# ------------------------------------------------
# Dictionary looping
d   = {  'a': [1, 2, 3] 
       , 'b': [2, 4, 6] 
       , 'c': [3, 5, 7]
        };      # Dictionary of lists
for k, v in d.items():
   print(k, v, 'list: ', v[0], v[1], v[2]);  # print each element


for k, v in d.items():
   print(d[k]);      # The dictionary of the key returns the value
   print(d[k][0]);   # Indexing into the returned value




# ------------------------------------------------
# Create a dictionary of lists from a list of lists 
l_list = [ [ 'a', 2, 'apple' ] 
          ,[ 'b', 4, 'bravo' ] 
          ,[ 'c', 5, 'crate' ] 
         ];     # list of lists - assume 1st element of each list is unique


d = {}; # Initialize an empty dictionary
for i in l_list: d[i[0]] = [int(i[1]), i[2]]; # Create the key i[0] and values list
print (d);


# Add a list to the dictionary
d['x'] = [11,'echo'];
print (d);


# Update a list in the dictionary
d['x'] = [22, 'golf'];
print (d);


# Delete a list from the dictionary
del(d['x']);
print (d);






#######################################################################################################################
## Lesson 07
# =======================
# Tuples
# =======================
t1 = (1, 2, 3, 4, 5);                 # tuple of integers
t2 = ('a', 1, 2.0, False);        # tuple of mixed data types
t3 = (1,);                      # One element tuple MUST have a comma
t4 = ();                           # Empty tuple
print(t1); print(t2); print(t3); print(t4);


# -------------
# tuple mutability (immutable - can't be modified in place) 
t = (1, 2, 3, 4, 5); # tuple of integers
print (t[0]);        # -> 1
t[0] = 10;           # -> Error


# -------------
# Operations
t = ('a', 1, 2.0, False);        # tuple of mixed data types
print("Length = ", len(t));         # Length
print("Is b in t? ", 'b' in (t));   # Membership                 


# Looping through a tuple
t = ('a', 1, 2.0, False);        # tuple of mixed data types


# Range Iteration
for i in range(0,len(t),1):     # Index Iteration
    print(t[i], end = ', ');  
print()


# Sequence Iteration
for i in (t):
    print(i, type(i));          # Sequence Iteration


# Slice 
print("Slice of element 1 and 2: ", t[1:3]);


# =============================================================
# Simple Lists
# =============================================================
l_list = [1, 2, 3, 4, 5, 6];
print(l_list, l_list[2], type(l_list[2]));  # Index a list
l_list [1] = 9; print(l_list);              # Change the 2nd element         


l_list = [1, 2, 3, 4, 5, 6];
del l_list[1];  # Delete an element using an index
print(l_list);         


# Remove the FIRST element from the list whose value is x. 
l_list = [1, 2, 3, 4, 3, 6, 4];
l_list.remove(4);        # remove the value 4 - not the element at index[4]
print(l_list); 


# Append an element to an existing list. 
l_list = [1, 2, 3, 4, 5, 6];
s = '7';
l_list.append(s);
print(l_list);




# =============================================================
# Nested sequences - List of Lists
# =============================================================
l_list = [ [ 1, 2, 3 ] 
          ,[ 2, 4, 6 ] 
          ,[ 3, 5, 7 ] 
         ];
print(l_list, type(l_list));


# Access a specific list, use the index :
print(l_list[0], type(l_list[0]));
# Access a specific list element:
print(l_list[0][2], type(l_list[0][2]));


# ADD a list to the list of lists:
i = int(input("Enter the first number of a list to add: "));
add_list = [ i, i*10, i-10 ];   # Add a list with [the entered number, number times 10, number - 10]
l_list.append(add_list);
print(l_list);


# DELETE a list from the list of lists using an index
for i in range(0, len(l_list), 1):    # Display the index associated with each element in the list
    print(i, ": ", l_list[i]);
i = int(input("Enter the index of an element to delete: "));   # User choose the index of the list
del l_list[i];  # Remove the list at the index position
print("\nNew list of lists: ", l_list);


# Update a list element in the list of lists using an index
l_list[2] = [ -10, -11, -12 ];
print(l_list);








#######################################################################################################################
## Lesson 06
# =======================
# for loops
# =======================
# for loops using the range function
for i in range(1,5,1):
   print(i, end=',');  # START =1, STOP =5  (not included), STEP =1


for i in range(1,5):
   print(i, end=',');  # START =1, STOP =5  (not included), STEP =1


for i in range(5):
   print(i, end=',');  # START =0, STOP =5  (not included), STEP =1


for i in range(1,10,2):
   print(i, end=',');  # Odd numbers


for i in range(5,0,-1):
   print(i, end=',');  # Backwards


# ----------------------
# for loops using a sequence
s = 'Python'; # string data type
for i in s:
   print(i);  # Each character in a string = Iterable sequence


# ----------------------
# Membership test on a sequence
# ----------------------
punctuation = '?, .!';  # string of punctuation characters
print ("In:  "   , '?' in     punctuation);
print ("Not In: ", 'y' not in punctuation);


2 in range(1,10,2);     # membership in a range of Odd numbers




# =======================
# Sequence Slicing 
# =======================
'''
  +---+---+---+---+---+---+
  | P | y | t | h | o | n |
  +---+---+---+---+---+---+
  0   1   2   3   4   5   6
'''
s = "Python";


print(s[0:6]);      # = 'Python'
print(s[0:len(s)]); # = 'Python'


print(s[0:2]);      # = 'Py'
print(s[:3]);       # First 3 characters = 'Pyt' 
print(s[4:]);       # Remaining characters after the h = 'on'
print(s[:3] + s[4:]);         # Original string with the ‘h’ removed 'Pyt' + 'on' = 'Pyton‘


# Special cases
print(s[:]);        # All characters
print(s[0:0]);      # End point = Start point = ''
print(s[3:1]);      # Illegal slice but no error = ''






#######################################################################################################################
## Lesson 05
# ----------------------
# While loop
# ----------------------
vi_count = 10;                  # Initialize loop variable
while vi_count > 0:             # Check loop variable
    vi_count = vi_count -1;     # Update loop variable
    print(vi_count, end=" ");
print("\n End of Count!");




# ----------------------
# Option Menu with while loop and conditional branching
# There is a logic error in the conditional branching for 'x'
# How should it be fixed?
# ----------------------
vs_choice = '';                 # Initialize the option choice loop variable
while vs_choice != 'x': # Check loop variable
    print("\nChoose a color for your new car"); 
    print(" 1. Red    ");
    print(" 2. Green  ");
    print(" 3. Yellow ");
    print(" 4. Blue   ");
    print(" x. Exit   ");


  # Get menu choice - Update loop variable
    vs_choice = input("\nEnter the number for your choice of color: "); 


    if   vs_choice == '1':
       print("You chose Red    ");
       
    elif vs_choice == '2':
       print("You chose Green  ");
       
    elif vs_choice == '3':
       print("You chose Yellow ");
       
    elif vs_choice == '4':
       print("You chose Blue   ");
       
    elif vs_choice == 'x':
       print("You chose Exit   ");
       
    else:
       print("\nThat is not a valid choice");


# End of while loop








# ----------------------
# Indexing
# ----------------------
'''
+---+---+---+---+---+---+
| P | y | t | h | o | n |
+---+---+---+---+---+---+
  0   1   2   3   4   5
 -6  -5  -4  -3  -2  -1 
'''
str = 'Python';  # Initialize string variable


# Print the string sequence: character by character forward
index = 0;                            # Initialize loop variable
while index < len(str):               # Check loop variable
     print (str[index], end = "-");   # print character at index
     index +=1;                       # Update loop variable
print();






#######################################################################################################################
## Lesson 04
# ----------------------
# Assignment 1 - Importing modules
# ----------------------
import random;
rand_nbr = random.randint(1, 6); print(rand_nbr); # randint 1-6 inclusive
rand_nbr = random.randrange(6);  print(rand_nbr); # randrange 0-5


from random import randint, randrange;
rand_nbr = randint(1, 6); print(rand_nbr); # randint 1-6 inclusive
rand_nbr = randrange(6);  print(rand_nbr); # randrange 0-5




# ----------------------
# Assignment 2 - Conditions and comparison operators
# ----------------------
a = 10; b = 20;
# False
print(a, b, a == b, a > b, a>=b);
# True
print(a, b, a!= b, a < b, a <= b);


# ---------
# Conditional Branching
# ---------
a = 10; b = 20;


# if 
if a != b:
    print(a, b);


# if\else
if a == b:
    print(a, b);
else:
    print("a not equal b");
    input("Press enter to continue");
    
i = 100;
# if\elif\else
if i >= 80:
    print("Grade = A"); # i must be >= 80
elif i >= 60:
    print("Grade = C"); # i must be < 80 AND >= 60
elif i < 60:
    print("Grade = F"); # i must be < 60
else:
    print("In else and not possible");  # This should not happen
    
# ---------
# Code blocks
# ---------
count = 2;


if (count == 3):
    count += 10;
    print("Count = ", count);   # Count is not printed since count not equal 3


if (count == 3):
    count += 10;
print("Count = ", count);






#######################################################################################################################
## Lesson 03
# ----------------------
# Numeric data type, operators, variables
# ----------------------
print(10);      # Integer
print(3.1417);  # Float


# Printing operations on the "fly"
print(10+20);   # Addition
print(7/3);     # True division
print(7//3);    # Integer division
print(7%3);     # Remainder after Integer division


x = 5;      print(x);   # Defining  x as 5
x = x +2;   print(x);   # Increment x by 2 
x +=3;      print(x);   # Increment x by 3 
y = x +5;   print(y);   # Defining y as addition to x


# data type() function
type(y);        # Integer
y = 8.1;        # Explicit re- assigning y as 8.1 and type float.
type(y);        # Float


z = 6;      print(z, type(z));  # Assigning z as 6 and type int
z = z//2;   print(z, type(z));  # Explicit re-assigning z as 3 and type int 
z = z/2;    print(z, type(z));  # Implicit redefine z as 3.0 and type float


b = True;   print(b, type(b));  # Assigning b as a boolean True
b = 5 < 3;  print(b, type(b));  # Re-Assigning b as a boolean 




# ----------------------
# Type conversion, String Methods, String Functions
# ----------------------
## Type conversion
vi = input("Enter number ");        # User input a number as a string
print(vi, type(vi));
vi = vi +1;   # Error


vi = int(input("Enter number "));   # Data type conversion on input
print(vi, type(vi));


## String Function
s = 'the quick brown fox';
ls = len(s);
print(s, '\t', ls, '\t',type(ls));  # String length


## String Method
us = s.upper();
print(s, '\t', us, '\t',type(us));  # Upper case


# ----------------------
# String Methods and Functions
# ----------------------
s = 'the quick brown fox';


# String Methods
s.capitalize(); # Returns the string with first letter capitalized and the rest lowercased.
s.count('o');   # Count the non-overlapping occurrence of supplied substring in the string.
s.find('o');    # Return the index of the first occurrence of supplied substring in the string. Return -1 if not found.
s.lower();      # Return a copy of all lowercased string.
s.replace('o','a'); # Replace all old substrings with new substrings.
s.split(' ');   # Function returns a list of all the words in the string, using str as the separator 
s.strip('x');   # Return a string with provided leading and trailing characters removed.
s.swapcase();   # Return a string with lowercase characters converted to uppercase and vice versa.
s.title();      # Return a title (first character of each word capitalized, others lowercased) cased string.
s.upper();      # Return a copy of all uppercased string.


# String Function
len(s); # Function Returns the length of the string (i.e. number of characters).




#######################################################################################################################
## Lesson 02
# ----------------------
# Printing String Operations
# ----------------------
print('');              # Empty string
print('1a2B3c4D5');     # Alphanumeric
print('!@\#$%^&*()_');  # Symbol characters including '\'


print('Program "Game Over" 2.0');       # Quotes inside strings
print('Mixed', "quotes", '''here''');   # Mixed quotes here


print("Same Line "); print("New line");                 # Two commands with semi-colon ending
print("Same Line", end = " -> "); print("Same Line");   # Using end = 


print("First  " +  "Second  " + " Third");  # Concatenation
print("More " *3);                # Repeating
print("line 1 " \
     + " line 2");                # Line continuation


print("\nNewline before");        # New line escape sequences


print('Tab:\t', '-> to here', '\nthen next line');  # Tab and New line after escape sequences
print('!line 1\n line 2 \\n lines');    # Symbol characters with embedded special code


# ----------------------
# String Variables
# ----------------------
lname = '';         # Creating variable lname as string data type
lname = 'Smith';    # Defining lname as Smith
print(lname);            # Display contents of variable lname
lname = 'Jones';    # Re-Defining lname as Jones 
print(lname);            # Display contents of variable lname


fname = '';         # Creating variable fname as string data type
print(fname);            # Display contents of variable fname


fname = 'Jane';            # Defining fname as Jane
print(fname, " ", lname);    # Display full name


# ----------------------
# Python Program
# ----------------------
name = "Larry";     # Assign a string to variable 'name'
print(name);        # Print the variable content
print("Hi,", name); # Print a string and variable 
print("\n");        # New line


name = input("Hi.  What's your name? ");    # Get name from user
print("Hi,\t", name);   # Print name input from user after TAB


input("\n\nPress the enter key to exit.");  # Exit program






#######################################################################################################################
## Lesson 01
# ----------------------
# Interactive mode
# ----------------------
print('Letters: A sequence of characters'); 
print("Numbers: 0 1 2 3 4 5 6 7 8 9 ");
print("My name is: ");  # Type your name after the colon


print(Error);   # Error - no quote
print('Error"); # Error - mismatch quotes
print "Error";  # Error - No Parens






# ----------------------
# Script mode
# ----------------------
#   File -> New File
#   File -> Save As -> find your directory -> name your file
'''
Multi-Line documentation
print and input fuctions
'''
print('This is a string');
input("Press any key to see the next print");


print('Numbers: 0 1 2 3 4 5 6 7 8 9');
input("Press the Enter key to exit");


#######################################################################################################################